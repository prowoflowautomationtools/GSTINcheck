# gstin_checker_app_v2.py
import streamlit as st
import pandas as pd
import requests
from datetime import datetime
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ---------------- Streamlit UI ----------------
st.set_page_config(page_title="GSTIN Checker", page_icon="🧾", layout="wide")

st.title("🧾 GSTIN Validation Tool")
st.markdown("""
Upload an Excel file containing a column named **GSTIN** in Column A (A1 header).  
Enter your **GSTIN Check API Key**, and the tool will fetch all details for each GSTIN and export an enriched Excel file.
""")

# Sidebar Info
with st.sidebar:
    st.header("Created by")
    st.write("**AuditFlow Team**")
    st.write("Version: 2.0")
    st.info("This version accepts the API key from the user for secure usage.")

# ---------------- Helper Function ----------------
def get_gstin_details(api_key, gstin):
    """Fetch GSTIN details from API"""
    url = f"http://sheet.gstincheck.co.in/check/{api_key}/{gstin}"
    try:
        r = requests.get(url, timeout=10)
        if r.status_code == 200:
            resp = r.json()
            if resp.get("flag"):
                data = resp.get("data", {})
                return {
                    "GSTIN": gstin,
                    "Status": data.get("sts", ""),
                    "Legal_Name": data.get("lgnm", ""),
                    "Trade_Name": data.get("tradeNam", ""),
                    "Constitution": data.get("ctb", ""),
                    "Center_Jurisdiction": data.get("ctj", ""),
                    "State_Jurisdiction": data.get("stj", ""),
                    "Principal_Place": data.get("pradr", {}).get("addr", ""),
                    "Registration_Date": data.get("rgdt", ""),
                    "Cancellation_Date": data.get("cxdt", ""),
                    "Nature_of_Business": ", ".join(data.get("nba", [])),
                    "State_Code": data.get("stcd", ""),
                    "Last_Updated": data.get("lstupdt", ""),
                }
            else:
                return {"GSTIN": gstin, "Status": "Invalid / Not Found"}
        else:
            return {"GSTIN": gstin, "Status": f"API Error {r.status_code}"}
    except Exception as e:
        return {"GSTIN": gstin, "Status": f"Error: {e}"}

# ---------------- Main UI ----------------
st.subheader("🔑 Enter your GSTIN Check API Key")
api_key = st.text_input("GSTIN API Key", type="password", help="Enter your API key from gstincheck.co.in")

uploaded_file = st.file_uploader("📤 Upload Excel file (must contain column 'GSTIN')", type=["xlsx", "xls"])

if uploaded_file and api_key:
    try:
        df = pd.read_excel(uploaded_file, engine="openpyxl")
    except Exception as e:
        st.error(f"Error reading Excel: {e}")
        st.stop()

    if "GSTIN" not in df.columns:
        st.error("❌ The uploaded file must have a column named 'GSTIN' in Column A.")
        st.stop()

    gstins = df["GSTIN"].dropna().astype(str).str.strip().unique().tolist()
    st.info(f"Found **{len(gstins)} unique GSTINs** for validation.")

    run_btn = st.button("🚀 Validate GSTINs")

    if run_btn:
        results = []
        progress = st.progress(0)
        status = st.empty()

        for idx, g in enumerate(gstins, start=1):
            status.info(f"Validating {g} ({idx}/{len(gstins)})")
            data = get_gstin_details(api_key, g)
            results.append(data)
            progress.progress(int(idx / len(gstins) * 100))

        status.success("✅ Validation completed successfully!")

        result_df = pd.DataFrame(results)

        st.subheader("📊 GSTIN Validation Results")
        st.dataframe(result_df)

        # Save Excel
        output_buffer = io.BytesIO()
        with pd.ExcelWriter(output_buffer, engine="openpyxl") as writer:
            result_df.to_excel(writer, index=False, sheet_name="GSTIN_Details")
            ws = writer.sheets["GSTIN_Details"]
            # Style header row
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for col in ws.columns:
                max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                ws.column_dimensions[col[0].column_letter].width = max_len + 2

        output_buffer.seek(0)
        file_name = f"GSTIN_Validation_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        st.download_button(
            label="📥 Download GSTIN Validation Report",
            data=output_buffer,
            file_name=file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

elif uploaded_file and not api_key:
    st.warning("⚠️ Please enter your API key before starting validation.")
